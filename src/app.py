import openpyxl
from tabulate import tabulate

# Define variable to load the dataframe
excel_dataframe = openpyxl.load_workbook("personas.xlsx")

# Define variable to read sheet
dataframe = excel_dataframe.active

data = []

# Iterate the loop to read the cell values
for row in range(1, dataframe.max_row):
    _row = [row,]

    for col in dataframe.iter_cols(1, dataframe.max_column):
        _row.append(col[row].value)

    data.append(_row)

headers = ["#", "Id", "Name", "Company", "Email", "MAC Address"]
headers_align = (("center",) * 6)

print(tabulate(data, headers=headers, tablefmt='fancy_grid', colalign=headers_align))
